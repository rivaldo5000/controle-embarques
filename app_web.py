import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import PatternFill

# Configuração da página Web
st.set_page_config(
    page_title="MVI • Gestão de Embarques",
    page_icon="🚚",
    layout="wide"
)

DB_NAME = "controle_embarques.db"

# Estilização CSS Personalizada para deixar o layout alegre e moderno
st.markdown("""
    <style>
        .stApp {
            background-color: #f8fafc;
        }
        .header-box {
            background-color: #006699;
            padding: 20px;
            border-radius: 12px;
            color: white;
            margin-bottom: 20px;
        }
        .kpi-card {
            background-color: white;
            padding: 15px;
            border-radius: 12px;
            border: 1px solid #e2e8f0;
            text-align: center;
            box-shadow: 0 2px 4px rgba(0,0,0,0.02);
        }
        .kpi-value {
            font-size: 24px;
            font-weight: bold;
        }
        .kpi-label {
            color: #64748b;
            font-size: 13px;
        }
    </style>
""", unsafe_allow_html=True)

# --- BANCO DE DADOS ---
def inicializar_banco():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS embarques (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cavalo TEXT,
            carreta TEXT,
            motorista TEXT,
            prev_chegada TEXT,
            loc_chegada TEXT,
            trava_chegada TEXT,
            chegada TEXT,
            status_prazo TEXT,
            prev_carreg TEXT,
            real_carreg TEXT,
            loc_carreg TEXT,
            trava_carreg TEXT,
            destino TEXT,
            frete TEXT,
            prazo TEXT
        )
    """)
    conn.commit()
    conn.close()

def converter_para_datetime(texto_data):
    if not texto_data or not str(texto_data).strip():
        return None
    
    texto_clean = str(texto_data).strip()
    formatos = [
        "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"
    ]
    for fmt in formatos:
        try:
            return datetime.strptime(texto_clean, fmt)
        except ValueError:
            continue
    return "ERRO"

def calcular_status_prazo(prev_str, real_str):
    if not real_str or not str(real_str).strip():
        return "Pendente"
    if not prev_str or not str(prev_str).strip():
        return "Sem Previsão"

    dt_prev = converter_para_datetime(prev_str)
    dt_real = converter_para_datetime(real_str)

    if dt_prev == "ERRO" or dt_real == "ERRO":
        return "Erro Formato"

    return "Dentro do Prazo" if dt_real <= dt_prev else "Atraso"

inicializar_banco()

# --- HEADER PRINCIPAL ---
st.markdown("""
    <div class="header-box">
        <h2 style="margin:0; color:white;">🚚 MVI • Gestão de Embarques & Chegadas</h2>
        <p style="margin:0; opacity:0.8; font-size:14px;">Painel de acompanhamento dinâmico e em tempo real</p>
    </div>
""", unsafe_allow_html=True)

# --- CARREGAR DADOS DOS CARDS (KPIs) ---
conn = sqlite3.connect(DB_NAME)
df_kpi = pd.read_sql_query("SELECT * FROM embarques", conn)
conn.close()

total_regs = len(df_kpi)
pendentes = len(df_kpi[df_kpi['chegada'].str.strip() == '']) if not df_kpi.empty else 0
concluidos = len(df_kpi[(df_kpi['chegada'].str.strip() != '') & (df_kpi['real_carreg'].str.strip() != '')]) if not df_kpi.empty else 0
atrasados = len(df_kpi[df_kpi['status_prazo'] == 'Atraso']) if not df_kpi.empty else 0

# Cards
k1, k2, k3, k4 = st.columns(4)
with k1:
    st.markdown(f'<div class="kpi-card"><div class="kpi-value" style="color:#0088cc;">{total_regs}</div><div class="kpi-label">Total Registros</div></div>', unsafe_allow_html=True)
with k2:
    st.markdown(f'<div class="kpi-card"><div class="kpi-value" style="color:#e63946;">{pendentes}</div><div class="kpi-label">Chegadas Pendentes</div></div>', unsafe_allow_html=True)
with k3:
    st.markdown(f'<div class="kpi-card"><div class="kpi-value" style="color:#2a9d8f;">{concluidos}</div><div class="kpi-label">Carregados (Verde)</div></div>', unsafe_allow_html=True)
with k4:
    st.markdown(f'<div class="kpi-card"><div class="kpi-value" style="color:#e76f51;">{atrasados}</div><div class="kpi-label">Com Atraso</div></div>', unsafe_allow_html=True)

st.write("")

# --- GERENCIAMENTO DE SELEÇÃO E EDIÇÃO ---
conn = sqlite3.connect(DB_NAME)
df_lista = pd.read_sql_query("SELECT id, cavalo, carreta, motorista FROM embarques ORDER BY id DESC", conn)
conn.close()

opcoes_edicao = ["➕ Novo Registro"] + [f"ID {row['id']} - Cavalo: {row['cavalo']} | Carreta: {row['carreta']}" for _, row in df_lista.iterrows()]

selecao = st.selectbox("📌 Selecione um registro existente para editar ou crie um novo:", opcoes_edicao)

id_selecionado = None
dados_atuais = {}

if selecao != "➕ Novo Registro":
    id_selecionado = int(selecao.split(" ")[1])
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM embarques WHERE id=?", (id_selecionado,))
    r = cursor.fetchone()
    conn.close()
    
    if r:
        dados_atuais = {
            "cavalo": r[1] or "", "carreta": r[2] or "", "motorista": r[3] or "", 
            "prev_chegada": r[4] or "", "loc_chegada": r[5] or "", "trava_chegada": r[6] or "", 
            "chegada": r[7] or "", "prev_carreg": r[9] or "", "real_carreg": r[10] or "", 
            "loc_carreg": r[11] or "", "trava_carreg": r[12] or "", "destino": r[13] or "SEST/VDC", 
            "frete": r[14] or "", "prazo": r[15] or ""
        }

# --- FORMULÁRIO DE CADASTRO / EDIÇÃO ---
with st.form("form_embarque", clear_on_submit=False):
    st.markdown(f"### {'✏️ Editando Registro ID ' + str(id_selecionado) if id_selecionado else '✨ Cadastro e Atualização de Veículos'}")
    
    c1, c2, c3, c4 = st.columns(4)
    cavalo = c1.text_input("Placa Cavalo*", value=dados_atuais.get("cavalo", "")).upper()
    carreta = c2.text_input("Placa Carreta", value=dados_atuais.get("carreta", "")).upper()
    motorista = c3.text_input("Motorista", value=dados_atuais.get("motorista", ""))
    frete = c4.text_input("Valor Frete (R$)", value=dados_atuais.get("frete", ""))

    c5, c6, c7, c8 = st.columns(4)
    prev_chegada = c5.text_input("Prev. Chegada", value=dados_atuais.get("prev_chegada", ""))
    loc_chegada = c6.text_input("Localizador (Chegada)", value=dados_atuais.get("loc_chegada", ""))
    trava_chegada = c7.text_input("Trava (Chegada)", value=dados_atuais.get("trava_chegada", ""))
    chegada_real = c8.text_input("Chegada Real", value=dados_atuais.get("chegada", ""))

    c9, c10, c11, c12 = st.columns(4)
    prev_carreg = c9.text_input("Prev. Carregamento", value=dados_atuais.get("prev_carreg", ""))
    real_carreg = c10.text_input("Data Real Carregamento", value=dados_atuais.get("real_carreg", ""))
    loc_carreg = c11.text_input("Localizador (Carreg.)", value=dados_atuais.get("loc_carreg", ""))
    trava_carreg = c12.text_input("Trava (Carreg.)", value=dados_atuais.get("trava_carreg", ""))

    c13, c14 = st.columns(2)
    dest_val = dados_atuais.get("destino", "SEST/VDC")
    idx_destino = 0 if dest_val == "SEST/VDC" else 1
    destino = c13.selectbox("Destino da Carga", ["SEST/VDC", "SEST/VDC/SP"], index=idx_destino)
    prazo = c14.text_input("Prazo Pagamento", value=dados_atuais.get("prazo", ""))

    salvar = st.form_submit_button("💾 Salvar Registro", type="primary")

    if salvar:
        if not cavalo:
            st.error("A Placa do Cavalo é obrigatória!")
        else:
            status_prazo = calcular_status_prazo(prev_chegada, chegada_real)
            conn = sqlite3.connect(DB_NAME)
            cursor = conn.cursor()

            if id_selecionado is None:
                cursor.execute("""
                    INSERT INTO embarques (
                        cavalo, carreta, motorista, prev_chegada, loc_chegada, trava_chegada,
                        chegada, status_prazo, prev_carreg, real_carreg, loc_carreg, trava_carreg,
                        destino, frete, prazo
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    cavalo, carreta, motorista, prev_chegada, loc_chegada, trava_chegada,
                    chegada_real, status_prazo, prev_carreg, real_carreg, loc_carreg, trava_carreg,
                    destino, frete, prazo
                ))
                st.success("Registro cadastrado com sucesso!")
            else:
                cursor.execute("""
                    UPDATE embarques SET
                        cavalo=?, carreta=?, motorista=?, prev_chegada=?, loc_chegada=?, trava_chegada=?,
                        chegada=?, status_prazo=?, prev_carreg=?, real_carreg=?, loc_carreg=?, trava_carreg=?,
                        destino=?, frete=?, prazo=?
                    WHERE id=?
                """, (
                    cavalo, carreta, motorista, prev_chegada, loc_chegada, trava_chegada,
                    chegada_real, status_prazo, prev_carreg, real_carreg, loc_carreg, trava_carreg,
                    destino, frete, prazo, id_selecionado
                ))
                st.success(f"Registro ID {id_selecionado} atualizado com sucesso!")

            conn.commit()
            conn.close()
            st.rerun()

if id_selecionado:
    if st.button(f"🗑️ Excluir Registro ID {id_selecionado}"):
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM embarques WHERE id=?", (id_selecionado,))
        conn.commit()
        conn.close()
        st.success(f"Registro ID {id_selecionado} excluído!")
        st.rerun()

st.divider()

# --- CONSULTA, FILTROS E TABELA ---
st.subheader("🔍 Filtros Rápidos")

f1, f2 = st.columns(2)
filtro_placa = f1.text_input("Filtrar por Placa (Cavalo ou Carreta)").strip().upper()
filtro_data = f2.text_input("Filtrar por Data (Carregamento/Chegada)").strip()

conn = sqlite3.connect(DB_NAME)
df = pd.read_sql_query("SELECT * FROM embarques ORDER BY id DESC", conn)
conn.close()

if not df.empty:
    if filtro_placa:
        df = df[df["cavalo"].str.contains(filtro_placa, na=False) | df["carreta"].str.contains(filtro_placa, na=False)]
    if filtro_data:
        df = df[df["real_carreg"].str.contains(filtro_data, na=False) | df["prev_carreg"].str.contains(filtro_data, na=False)]

    def aplicar_cores(row):
        chegada = row["chegada"]
        real_carreg = row["real_carreg"]
        
        cor = ""
        if not chegada:
            cor = "background-color: #fecdd3; color: #881337;" # Vermelho Pastel
        elif chegada:
            cor = "background-color: #e0f2fe; color: #0369a1;" # Azul Pastel

        if not real_carreg and chegada:
            cor = "background-color: #fef08a; color: #713f12;" # Amarelo Pastel
        elif real_carreg and chegada:
            cor = "background-color: #dcfce7; color: #14532d;" # Verde Pastel

        return [cor] * len(row)

    st.dataframe(df.style.apply(aplicar_cores, axis=1), use_container_width=True, height=400)

    # --- EXPORTAÇÃO EXCEL ---
    def gerar_excel(data_frame):
        output = io.BytesIO()
        df_export = data_frame.copy()
        
        df_export.rename(columns={
            "id": "ID", "cavalo": "Placa Cavalo", "carreta": "Placa Carreta", 
            "motorista": "Motorista", "prev_chegada": "Prev. Chegada", 
            "loc_chegada": "Localizador (Chegada)", "trava_chegada": "Trava (Chegada)",
            "chegada": "Chegada Real", "status_prazo": "Status Prazo", 
            "prev_carreg": "Prev. Carregamento", "real_carreg": "Data Real Carregamento", 
            "loc_carreg": "Localizador (Carregamento)", "trava_carreg": "Trava (Carregamento)",
            "destino": "Destino", "frete": "Valor Frete", "prazo": "Prazo Pagamento"
        }, inplace=True)

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Embarques')
        
        output.seek(0)
        wb = openpyxl.load_workbook(output)
        ws = wb.active

        fill_vermelho = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        fill_azul = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        fill_amarelo = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        fill_verde = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            chegada_val = ws.cell(row=row, column=8).value
            real_carreg_val = ws.cell(row=row, column=11).value

            fill_aplicar = None
            if not chegada_val:
                fill_aplicar = fill_vermelho
            elif chegada_val:
                fill_aplicar = fill_azul

            if not real_carreg_val and chegada_val:
                fill_aplicar = fill_amarelo
            elif real_carreg_val and chegada_val:
                fill_aplicar = fill_verde

            if fill_aplicar:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill_aplicar

        out_final = io.BytesIO()
        wb.save(out_final)
        out_final.seek(0)
        return out_final

    excel_data = gerar_excel(df)
    st.download_button(
        label="📊 Exportar Relatório Formatado em Excel (.xlsx)",
        data=excel_data,
        file_name=f"relatorio_embarques_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
else:
    st.info("Nenhum registro encontrado no banco de dados.")
